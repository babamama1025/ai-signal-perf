"""openpyxl Excel 匯出模組"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# 色彩定義
FILL_HEADER = PatternFill('solid', fgColor='4472C4')
FILL_SUBHEADER = PatternFill('solid', fgColor='9DC3E6')
FILL_GREEN = PatternFill('solid', fgColor='C6EFCE')
FILL_RED = PatternFill('solid', fgColor='FFC7CE')
FILL_SUMMARY_HDR = PatternFill('solid', fgColor='D9E1F2')

FONT_WHITE_BOLD = Font(name='微軟正黑體', bold=True, color='FFFFFF')
FONT_BOLD = Font(name='微軟正黑體', bold=True)
FONT_GREEN = Font(name='微軟正黑體', color='375623')
FONT_RED = Font(name='微軟正黑體', color='9C0006')
FONT_NORMAL = Font(name='微軟正黑體')

from comparison_logic import LOWER_BETTER, METRIC_UNITS, aggregate_periods


def build_comparison_xlsx(
    all_results: dict[str, dict[str, pd.DataFrame]],
    before_by_period: dict[str, list],
    after_by_period: dict[str, list],
    include_travel_time: bool = True,
    raw_df: pd.DataFrame | None = None,
    raw_periods: list | None = None,
    extra_summary_entities: list[str] | None = None,
) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    _build_info_sheet(wb, before_by_period, after_by_period)
    _build_summary_sheet(wb, all_results, before_by_period, after_by_period,
                         extra_summary_entities=extra_summary_entities)
    for period, results in all_results.items():
        _build_period_sheet(
            wb, period, results,
            before_by_period.get(period, []), after_by_period.get(period, []),
            include_travel_time,
        )
    if raw_df is not None and raw_periods:
        _build_raw_data_sheet(wb, raw_df, before_by_period, after_by_period, raw_periods)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fmt_date(d) -> str:
    _WD = ['一', '二', '三', '四', '五', '六', '日']
    ts = d if isinstance(d, pd.Timestamp) else pd.Timestamp(d)
    return f"{ts.strftime('%Y/%m/%d')} (週{_WD[ts.weekday()]})"


def _build_info_sheet(wb, before_by_period, after_by_period):
    from datetime import datetime
    ws = wb.create_sheet('分析說明')

    ws.merge_cells('A1:C1')
    ws['A1'] = 'AI 號誌績效比較分析'
    ws['A1'].font = Font(name='微軟正黑體', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f"產製時間：{datetime.now().strftime('%Y/%m/%d %H:%M')}"
    ws['A2'].font = Font(name='微軟正黑體', italic=True, color='666666')

    row = 4
    for period in before_by_period:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row, 1, f'⏱ {period}')
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_HEADER
        row += 1
        for label, dates, fill in [
            (f'事前日期（定時時制，共 {len(before_by_period[period])} 天）', before_by_period[period], FILL_SUBHEADER),
            (f'事後日期（AI 號誌，共 {len(after_by_period[period])} 天）',   after_by_period[period],  FILL_SUMMARY_HDR),
        ]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell = ws.cell(row, 1, label)
            cell.font = FONT_BOLD
            cell.fill = fill
            row += 1
            for d in dates:
                ws.cell(row, 1, _fmt_date(d)).font = FONT_NORMAL
                row += 1
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 28


def _build_summary_sheet(wb, all_results, before_by_period, after_by_period,
                          extra_summary_entities: list[str] | None = None):
    ws = wb.create_sheet('總表')
    periods = list(all_results.keys())
    extra_entities = list(extra_summary_entities) if extra_summary_entities else []

    # 「全時段合計」：總停等延滯／通過量加總、平均停等延滯由加總重新推導（不含旅行時間）
    aggregated = aggregate_periods(all_results, periods, include_travel_time=False,
                                   extra_entities=extra_entities or None)
    n_before_all = sum(len(before_by_period.get(p, [])) for p in periods)
    n_after_all  = sum(len(after_by_period.get(p, [])) for p in periods)

    col_blocks = [('全時段合計', aggregated, n_before_all, n_after_all)]
    col_blocks += [
        (period, all_results[period], len(before_by_period.get(period, [])), len(after_by_period.get(period, [])))
        for period in periods
    ]

    # 標題列
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(1, 1, '指標').font = FONT_BOLD
    ws.cell(1, 1).fill = FILL_SUMMARY_HDR
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    col_start = 2
    sub_headers = ['事前平均', '事後平均', '差異', '改善%']
    for i, (label, _, n_before, n_after) in enumerate(col_blocks):
        c = col_start + i * 4
        ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + 3)
        ws.cell(1, c, f'{label}（前{n_before}/後{n_after}）').font = FONT_BOLD
        ws.cell(1, c).fill = FILL_SUMMARY_HDR
        ws.cell(1, c).alignment = Alignment(horizontal='center')
        for j, h in enumerate(sub_headers):
            ws.cell(2, c + j, h).font = FONT_BOLD
            ws.cell(2, c + j).fill = FILL_SUMMARY_HDR

    row = 3
    first_period = periods[0]
    metrics_in_results = [k for k in all_results[first_period] if k != '旅行時間']
    for metric in metrics_in_results:
        # 指標標題列（顯示系統層級數據）
        ws.cell(row, 1, metric).font = FONT_WHITE_BOLD
        ws.cell(row, 1).fill = FILL_HEADER
        for i, (_, results_dict, _n_before, _n_after) in enumerate(col_blocks):
            c = col_start + i * 4
            df = results_dict.get(metric, pd.DataFrame())
            r = df[df['欄位'] == '系統'] if not df.empty else pd.DataFrame()
            if not r.empty:
                b, a = r['事前平均'].values[0], r['事後平均'].values[0]
                diff, pct = r['差異'].values[0], r['改善%'].values[0]
                ws.cell(row, c, b if not pd.isna(b) else None)
                ws.cell(row, c + 1, a if not pd.isna(a) else None)
                ws.cell(row, c + 2, diff if not pd.isna(diff) else None)
                _apply_num_format(ws.cell(row, c), metric)
                _apply_num_format(ws.cell(row, c + 1), metric)
                _apply_num_format(ws.cell(row, c + 2), metric)
                _write_pct(ws.cell(row, c + 3), pct)
        row += 1

        # 額外實體子列（桃園三期：高鐵周邊、A19周邊、前期範圍）
        for entity in extra_entities:
            ws.cell(row, 1, f'  {entity}').font = FONT_NORMAL
            ws.cell(row, 1).fill = FILL_SUBHEADER
            for i, (_, results_dict, _n_before, _n_after) in enumerate(col_blocks):
                c = col_start + i * 4
                df = results_dict.get(metric, pd.DataFrame())
                r = df[df['欄位'] == entity] if not df.empty else pd.DataFrame()
                if not r.empty:
                    b, a = r['事前平均'].values[0], r['事後平均'].values[0]
                    diff, pct = r['差異'].values[0], r['改善%'].values[0]
                    ws.cell(row, c, b if not pd.isna(b) else None)
                    ws.cell(row, c + 1, a if not pd.isna(a) else None)
                    ws.cell(row, c + 2, diff if not pd.isna(diff) else None)
                    _apply_num_format(ws.cell(row, c), metric)
                    _apply_num_format(ws.cell(row, c + 1), metric)
                    _apply_num_format(ws.cell(row, c + 2), metric)
                    _write_pct(ws.cell(row, c + 3), pct)
            row += 1

    # 欄寬
    ws.column_dimensions['A'].width = 18
    for col in range(2, 2 + len(col_blocks) * 4):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # 說明（各時段事前／事後天數已列於欄標題）
    ws.cell(row + 1, 1, '「全時段合計」為所有時段加總（旅行時間類指標除外）；各時段事前／事後天數詳見欄標題；完整日期清單見「分析說明」工作表').font = Font(italic=True, color='666666')


def _build_period_sheet(wb, period, results, before_dates, after_dates, include_travel_time):
    safe_name = period.replace(':', '').replace('~', '-')[:31]
    ws = wb.create_sheet(safe_name)

    # 依指標順序並排顯示（排除旅行時間；旅行時間接在最後一個路口來向列之後）
    metrics_in_results = [k for k in results if k != '旅行時間']
    n_metrics = len(metrics_in_results)
    total_cols = max(1 + n_metrics * 4, 5)

    # 第一列：時段標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1, f"{period}  │  事前：{len(before_dates)} 日  │  事後：{len(after_dates)} 日")
    ws.cell(1, 1).font = Font(name='微軟正黑體', bold=True, size=12)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # 第二、三列：欄位標頭（各指標並排：事前平均／事後平均／差異／改善%）
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    item_cell = ws.cell(2, 1, '項目')
    item_cell.font = FONT_WHITE_BOLD
    item_cell.fill = FILL_HEADER
    item_cell.alignment = Alignment(horizontal='center', vertical='center')

    sub_headers = ['事前平均', '事後平均', '差異', '改善%']
    for m, metric in enumerate(metrics_in_results):
        c = 2 + m * 4
        unit = METRIC_UNITS.get(metric, '')
        ws.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c + 3)
        head_cell = ws.cell(2, c, f'{metric} ({unit})')
        head_cell.font = FONT_WHITE_BOLD
        head_cell.fill = FILL_HEADER
        head_cell.alignment = Alignment(horizontal='center')
        for i, h in enumerate(sub_headers):
            sub_cell = ws.cell(3, c + i, h)
            sub_cell.font = FONT_WHITE_BOLD
            sub_cell.fill = FILL_HEADER
            sub_cell.alignment = Alignment(horizontal='center')

    # 資料列：欄位順序取各指標資料的聯集（依出現順序）
    field_order = []
    seen = set()
    for metric in metrics_in_results:
        mdf = results.get(metric, pd.DataFrame())
        if mdf.empty:
            continue
        for f in mdf['欄位']:
            if f not in seen:
                seen.add(f)
                field_order.append(f)

    row = 4
    for field in field_order:
        ws.cell(row, 1, field).font = FONT_NORMAL
        for m, metric in enumerate(metrics_in_results):
            c = 2 + m * 4
            mdf = results.get(metric, pd.DataFrame())
            r = mdf[mdf['欄位'] == field] if not mdf.empty else pd.DataFrame()
            if not r.empty:
                b, a, diff, pct = (r[col].values[0] for col in ['事前平均', '事後平均', '差異', '改善%'])
                _write_num(ws.cell(row, c), b, metric)
                _write_num(ws.cell(row, c + 1), a, metric)
                _write_num(ws.cell(row, c + 2), diff, metric)
                _write_pct(ws.cell(row, c + 3), pct)
            for cc in range(c, c + 4):
                ws.cell(row, cc).alignment = Alignment(horizontal='right')
        row += 1

    # 旅行時間：接續在最後一個路口來向列之後
    tt_df = results.get('旅行時間', pd.DataFrame())
    if include_travel_time and not tt_df.empty:
        row = _write_section(
            ws, row, f'▶ 旅行時間 ({METRIC_UNITS.get("旅行時間", "")})',
            tt_df, '旅行時間', FILL_HEADER, total_cols,
        )

    # 欄寬
    ws.column_dimensions['A'].width = 30
    for col in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # 凍結前三列
    ws.freeze_panes = 'A4'


def _build_raw_data_sheet(wb, df: pd.DataFrame, before_by_period, after_by_period, periods):
    ws = wb.create_sheet('原始資料')

    parts = []
    for period in periods:
        before_set = {pd.Timestamp(d) for d in before_by_period.get(period, [])}
        after_set  = {pd.Timestamp(d) for d in after_by_period.get(period, [])}
        date_set   = before_set | after_set
        sub = df[(df['時段'] == period) & (df['日期'].isin(date_set))].copy()
        sub.insert(0, '分組', sub['日期'].apply(
            lambda d: '事前' if d in before_set else '事後'
        ))
        parts.append(sub)
    raw = pd.concat(parts, ignore_index=True) if parts else df.iloc[0:0].copy()
    raw['日期'] = raw['日期'].apply(lambda d: d.strftime('%Y/%m/%d'))

    headers = raw.columns.tolist()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal='center')

    for r, (_, row_data) in enumerate(raw.iterrows(), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r, c)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, float) and val == int(val):
                cell.value = int(val)
            else:
                cell.value = val
            cell.font = FONT_NORMAL

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    for i in range(5, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.freeze_panes = 'A2'


def _write_section(ws, row: int, header: str, df: pd.DataFrame, metric: str, fill, total_cols: int = 5) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, header)
    if fill == FILL_HEADER:
        cell.font = FONT_WHITE_BOLD
    else:
        cell.font = FONT_BOLD
    cell.fill = fill
    row += 1

    if df.empty:
        ws.cell(row, 1, '（無資料）').font = Font(italic=True, color='888888')
        return row + 1

    for _, data_row in df.iterrows():
        field = data_row['欄位']
        b = data_row['事前平均']
        a = data_row['事後平均']
        diff = data_row['差異']
        pct = data_row['改善%']

        ws.cell(row, 1, field).font = FONT_NORMAL
        _write_num(ws.cell(row, 2), b, metric)
        _write_num(ws.cell(row, 3), a, metric)
        _write_num(ws.cell(row, 4), diff, metric)

        _write_pct(ws.cell(row, 5), pct)

        for c in range(2, 6):
            ws.cell(row, c).alignment = Alignment(horizontal='right')
        row += 1

    return row


def _write_pct(cell, pct):
    if not pd.isna(pct):
        cell.value = pct
        cell.number_format = '+0.0%;-0.0%;0.0%'
        if pct > 0:
            cell.fill = FILL_GREEN
            cell.font = FONT_GREEN
        elif pct < 0:
            cell.fill = FILL_RED
            cell.font = FONT_RED
    else:
        cell.value = '—'


def _write_num(cell, val, metric: str):
    if pd.isna(val):
        cell.value = '—'
        return
    cell.value = val
    _apply_num_format(cell, metric)


def _apply_num_format(cell, metric: str):
    # LOWER_BETTER 指標為小數格式；非 LOWER_BETTER（通過量類）為整數格式
    cell.number_format = '#,##0' if metric not in LOWER_BETTER else '#,##0.0'
    cell.font = FONT_NORMAL
